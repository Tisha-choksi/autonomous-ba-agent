import pandas as pd
import numpy as np
from pathlib import Path
from datetime import datetime
import json
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.colors import HexColor, white, black
from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Table,
                                  TableStyle, HRFlowable, Image as RLImage, PageBreak)
from reportlab.lib.units import cm, mm
from reportlab.lib import colors
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io
import base64
from .load_data import get_dataframe
from .eda import run_eda
from .kpi import calculate_kpis
from .insights import generate_insights

EXPORTS_PATH = Path("./exports")
EXPORTS_PATH.mkdir(parents=True, exist_ok=True)

BRAND_DARK = HexColor("#0F172A")
BRAND_BLUE = HexColor("#2563EB")
BRAND_SLATE = HexColor("#1E293B")
BRAND_LIGHT = HexColor("#F1F5F9")

def generate_pdf_report(session_id: str, file_name: str = None) -> dict:
    df = get_dataframe(session_id)
    if df is None:
        return {"error": "No data loaded"}

    eda = run_eda(session_id)
    kpis = calculate_kpis(session_id)
    insights = generate_insights(session_id)

    out_name = file_name or f"ba_report_{session_id[:8]}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    out_path = EXPORTS_PATH / out_name

    doc = SimpleDocTemplate(str(out_path), pagesize=A4,
                            leftMargin=2*cm, rightMargin=2*cm,
                            topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle("Title", parent=styles["Normal"],
        fontSize=22, textColor=BRAND_BLUE, spaceAfter=6, fontName="Helvetica-Bold")
    h2_style = ParagraphStyle("H2", parent=styles["Normal"],
        fontSize=14, textColor=BRAND_BLUE, spaceBefore=14, spaceAfter=6, fontName="Helvetica-Bold")
    body_style = ParagraphStyle("Body", parent=styles["Normal"],
        fontSize=10, textColor=HexColor("#334155"), spaceAfter=4, fontName="Helvetica")

    story = []

    # Header
    story.append(Paragraph("📊 Business Analysis Report", title_style))
    story.append(Paragraph(f"Generated: {datetime.now().strftime('%B %d, %Y %H:%M')} | Session: {session_id[:12]}", body_style))
    story.append(HRFlowable(width="100%", thickness=2, color=BRAND_BLUE))
    story.append(Spacer(1, 0.4*cm))

    # Dataset Overview
    story.append(Paragraph("Dataset Overview", h2_style))
    overview_data = [
        ["Metric", "Value"],
        ["Total Rows", f"{eda['shape']['rows']:,}"],
        ["Total Columns", str(eda['shape']['columns'])],
        ["Numeric Columns", str(len(eda['columns']['numeric']))],
        ["Categorical Columns", str(len(eda['columns']['categorical']))],
        ["Missing Values", str(len(eda.get('missing_values', {}))) + " columns affected"],
        ["Duplicate Rows", str(eda.get('duplicates', 0))],
    ]
    t = Table(overview_data, colWidths=[8*cm, 8*cm])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), BRAND_BLUE),
        ("TEXTCOLOR", (0,0), (-1,0), white),
        ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE", (0,0), (-1,0), 11),
        ("BACKGROUND", (0,1), (-1,-1), HexColor("#F8FAFC")),
        ("ROWBACKGROUNDS", (0,1), (-1,-1), [white, HexColor("#F1F5F9")]),
        ("GRID", (0,0), (-1,-1), 0.5, HexColor("#CBD5E1")),
        ("FONTSIZE", (0,1), (-1,-1), 10),
        ("PADDING", (0,0), (-1,-1), 8),
    ]))
    story.append(t)
    story.append(Spacer(1, 0.4*cm))

    # KPIs
    if kpis:
        story.append(Paragraph("Key Performance Indicators", h2_style))
        kpi_rows = [["KPI", "Value"]]
        for k, v in list(kpis.items())[:12]:
            val = v.get("value", 0)
            fmt = v.get("format", "number")
            if fmt == "currency": display = f"${val:,.2f}"
            elif fmt == "percent": display = f"{val:.1f}%"
            else: display = f"{val:,.0f}" if isinstance(val, (int, float)) else str(val)
            kpi_rows.append([v.get("label", k), display])
        kt = Table(kpi_rows, colWidths=[10*cm, 6*cm])
        kt.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,0), BRAND_BLUE),
            ("TEXTCOLOR", (0,0), (-1,0), white),
            ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
            ("ROWBACKGROUNDS", (0,1), (-1,-1), [white, HexColor("#F1F5F9")]),
            ("GRID", (0,0), (-1,-1), 0.5, HexColor("#CBD5E1")),
            ("FONTSIZE", (0,0), (-1,-1), 10),
            ("PADDING", (0,0), (-1,-1), 8),
        ]))
        story.append(kt)
        story.append(Spacer(1, 0.4*cm))

    # Insights
    if insights:
        story.append(Paragraph("AI-Generated Insights", h2_style))
        for i, insight in enumerate(insights[:8], 1):
            severity_color = {"warning": "#F59E0B", "success": "#10B981", "info": "#3B82F6"}.get(
                insight.get("severity", "info"), "#3B82F6")
            story.append(Paragraph(f"<b>{i}. {insight['title']}</b>", 
                ParagraphStyle("ins", parent=body_style, textColor=HexColor(severity_color), fontSize=11)))
            story.append(Paragraph(insight["description"], body_style))
            story.append(Spacer(1, 0.2*cm))

    # Data Sample
    story.append(PageBreak())
    story.append(Paragraph("Data Sample (First 20 Rows)", h2_style))
    sample = df.head(20)
    cols = sample.columns.tolist()[:8]  # max 8 cols for PDF width
    table_data = [[str(c)[:15] for c in cols]]
    for _, row in sample[cols].iterrows():
        table_data.append([str(row[c])[:15] if pd.notna(row[c]) else "" for c in cols])

    col_w = 17 * cm / len(cols)
    dt = Table(table_data, colWidths=[col_w] * len(cols))
    dt.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), BRAND_BLUE),
        ("TEXTCOLOR", (0,0), (-1,0), white),
        ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
        ("ROWBACKGROUNDS", (0,1), (-1,-1), [white, HexColor("#F1F5F9")]),
        ("GRID", (0,0), (-1,-1), 0.3, HexColor("#E2E8F0")),
        ("FONTSIZE", (0,0), (-1,-1), 8),
        ("PADDING", (0,0), (-1,-1), 5),
    ]))
    story.append(dt)

    doc.build(story)
    return {"file_path": str(out_path), "file_name": out_name, "report_type": "pdf"}

def generate_excel_report(session_id: str, file_name: str = None) -> dict:
    df = get_dataframe(session_id)
    if df is None:
        return {"error": "No data loaded"}

    eda = run_eda(session_id)
    kpis = calculate_kpis(session_id)
    insights = generate_insights(session_id)

    out_name = file_name or f"ba_report_{session_id[:8]}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    out_path = EXPORTS_PATH / out_name

    wb = openpyxl.Workbook()

    # Style helpers
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="2563EB")
    alt_fill = PatternFill("solid", fgColor="F1F5F9")
    border = Border(
        bottom=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1")
    )

    def style_header_row(ws, row_num, n_cols):
        for c in range(1, n_cols + 1):
            cell = ws.cell(row=row_num, column=c)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

    def auto_col_width(ws):
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=10)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)

    # Sheet 1: Raw Data
    ws1 = wb.active
    ws1.title = "📋 Data"
    ws1.append(df.columns.tolist())
    style_header_row(ws1, 1, len(df.columns))
    for i, (_, row) in enumerate(df.iterrows()):
        ws1.append([str(v) if pd.notna(v) else "" for v in row.values])
        if i % 2 == 1:
            for c in range(1, len(df.columns) + 1):
                ws1.cell(row=i+2, column=c).fill = alt_fill
    auto_col_width(ws1)

    # Sheet 2: EDA Summary
    ws2 = wb.create_sheet("📊 EDA Summary")
    ws2.append(["Metric", "Value"])
    style_header_row(ws2, 1, 2)
    ws2.append(["Total Rows", eda['shape']['rows']])
    ws2.append(["Total Columns", eda['shape']['columns']])
    ws2.append(["Numeric Columns", len(eda['columns']['numeric'])])
    ws2.append(["Categorical Columns", len(eda['columns']['categorical'])])
    ws2.append(["Duplicate Rows", eda.get('duplicates', 0)])
    if eda.get("stats", {}).get("numeric_summary"):
        ws2.append([])
        ws2.append(["Column Statistics"])
        for col, stats in eda["stats"]["numeric_summary"].items():
            ws2.append([f"--- {col} ---"])
            for k, v in stats.items():
                ws2.append([k, v])
    auto_col_width(ws2)

    # Sheet 3: KPIs
    ws3 = wb.create_sheet("🎯 KPIs")
    ws3.append(["KPI", "Value", "Format"])
    style_header_row(ws3, 1, 3)
    for k, v in kpis.items():
        ws3.append([v.get("label", k), v.get("value"), v.get("format")])
    auto_col_width(ws3)

    # Sheet 4: Insights
    ws4 = wb.create_sheet("💡 Insights")
    ws4.append(["#", "Type", "Title", "Description", "Severity"])
    style_header_row(ws4, 1, 5)
    for i, ins in enumerate(insights, 1):
        ws4.append([i, ins.get("type",""), ins.get("title",""),
                    ins.get("description",""), ins.get("severity","")])
    auto_col_width(ws4)

    wb.save(str(out_path))
    return {"file_path": str(out_path), "file_name": out_name, "report_type": "excel"}